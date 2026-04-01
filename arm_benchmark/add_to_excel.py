import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def add_value_to_excel(excel_file: str, column_num: int, value: str) -> None:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    column_letter = get_column_letter(column_num)
    latest_row = sheet.max_row

    target = sheet[f"{column_letter}{latest_row}"]
    if target.value is None:
      target.value = value
    else:
      sheet[f"{column_letter}{latest_row + 1}"] = value

    workbook.save(excel_file)


def main() -> int:
    if len(sys.argv) != 4:
        print("Usage: python add_to_excel.py <excel_file> <column_number> <value>")
        return 1

    excel_file = Path(sys.argv[1])
    if not excel_file.exists():
        print(f"Excel file not found: {excel_file}")
        return 1

    try:
        column_num = int(sys.argv[2])
    except ValueError:
        print("column_number must be an integer")
        return 1

    add_value_to_excel(str(excel_file), column_num, sys.argv[3])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
